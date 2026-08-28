#!/usr/bin/env python3
"""Regenerate PARTS-TRACKER.xlsx. Edit ROWS below, re-run, re-verify.
Found-column data comes from the user's store carts (2026-08 shopping)."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
def F(**kw):
    base = {"name": ARIAL, "size": 10}
    base.update(kw)
    return Font(**base)

YELLOW = PatternFill("solid", fgColor="FFF2AC")
SECTION = PatternFill("solid", fgColor="D9D6CB")
HEADER = PatternFill("solid", fgColor="444441")
EXAMPLE = Font(name=ARIAL, size=10, italic=True, color="8C919C")
thin = Side(style="thin", color="B4B2A9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '$#,##0.00'

# tag, item, qty, spec, est, suggested, FOUND-AT, ACTUAL, bought
ROWS = [
 ("EX", "(example row — overwrite or ignore)", "", "shows the format", None, "", "Micro Center Marietta", 42.99, "Aug 23"),
 ("SEC", "S1 — FIRST LIGHT (one panel, bonnet)", None),
 ("S1", "Raspberry Pi 5 (1 GB or 2 GB)", 1, "1 GB works (zram configured); cut Amazon SANOOV 2 GB kit $109.99 dupe", 43.00, "Micro Center", "MC SKU 959668", 42.99, "Aug 21"),
 ("S1", "microSD 32 GB+", 1, "32 GB is plenty — cut Amazon SanDisk 64 GB $25.99 dupe; keep an image backup", 10.00, "Micro Center", "MC SKU 482083", 9.99, "Aug 21"),
 ("S1", "USB-C PD charger 27 W+", 1, "a charger you OWN is still $0; cut Amazon official-27W $23.91 dupe", 0.00, "drawer / Micro Center", "MC SKU 620351", 11.99, "Aug 21"),
 ("S1", "Pi 5 Active Cooler", 1, "", 6.00, "Micro Center", "MC SKU 620310", 9.99, "Aug 21"),
 ("S1", "Adafruit Triple LED Matrix Bonnet + riser", 1, "CHECK THE RISER BOX on the product page (+$1.95) — that kills the Amazon Frienda 8-pk $9.99 and MC stacking kit $13.99", 12.00, "Adafruit", "Adafruit #6358 + riser add-on", 11.90, ""),
 ("S1", "Mean Well LRS-50-5 (5 V/10 A)", 1, "authorized distributor only — Amazon 3P listing ($16.65 shipped) is not one, and Mouser is cheaper anyway", 16.00, "Mouser", "Mouser #709-LRS-50-5", None, ""),
 ("S1", "IEC cord (becomes mains pigtail)", 1, "Amazon Basics 2-pk $11.18 beats MC single $12.99 — cut MC cord; cord #2 lands in S2", 0.00, "Amazon", "Amazon Basics C13 6 ft (x2 pack, 1 of 2)", 5.59, ""),
 ("S1", "TCS34725 color sensor", 1, "2-pk $8.99 beats Adafruit #1334 $7.95 single (spare incl); LED stays off during reads — script handles", 8.00, "Amazon", "Amazon GODIYMODULES 2-pk", 8.99, ""),
 ("SEC", "S2 — THE 9-PANEL WALL", None),
 ("S2", "P2.5 64x64 HUB75 panels x10", 10, "Pi Hut Waveshare RGB-Matrix-P2.5-64x64 is the pick: keeps the 480 mm plan, page confirms 1/32 scan ABCDE + synchronous driver + <=20 W (9x20=180 W on the 300 W PSU), and a reviewer runs it on the Adafruit bonnet. Buy 10 = 9 + same-batch spare. CHECK: does the $325 include US import duty? (China origin, de minimis is gone — carrier may bill after.) AliExpress ~$200 stays the gamble option; Amazon P3 $376 = out", 252.00, "The Pi Hut (Waveshare)", "Pi Hut WAV-23708, GBP 20.83 ex-VAT ea", 325.00, "Aug 21"),
 ("S2", "Mean Well LRS-350-5 (5 V/60 A)", 1, "same rule — Amazon 3P $37.04 shipped vs Mouser ~$35 w/ free ship over $50 order", 45.00, "Mouser", "Mouser #709-LRS-350-5", None, ""),
 ("S2", "SL22 10005 NTC inrush limiter", 1, "AMAZON CART HAS THE WRONG PART: SL22 0R516 is 0.5 ohm (limits nothing here; the sim used 10 ohm) + $9.90 Radwell shipping on a $1.82 part. Get 10005 in the Mouser order", 3.00, "Mouser", "Mouser #995-SL2210005", None, ""),
 ("S2", "14 AWG silicone wire, red+black", 1, "25 ft 2-core spool = 25 ft each color; covers trunk + drops", 12.00, "Amazon", "Amazon Haerkn 2-core 25 ft", 19.98, ""),
 ("S2", "Bus bars / power distribution blocks", 2, "pair (pos+neg), 150 A, covers + lugs incl; skip if you fab the backplane PCB instead", 10.00, "Amazon", "Amazon RVBOATPAT pair", 15.99, ""),
 ("S2", "Inline ATC fuse holders, 14 AWG", 10, "FUSES NOT INCLUDED — next row; skip if fabbing the PCB (it has fuse clips)", 12.00, "Amazon", "Amazon Nilight 10-pk", 9.55, ""),
 ("S2", "ATO blade fuses 7.5 A", 1, "100-pc + pullers; 7.5 A per injection drop is right for 2 panels/drop", 6.00, "Amazon", "Amazon Vigor_Source 100-pc", 5.99, ""),
 ("S2", "Fork/ring terminal assortment", 1, "heat-shrink insulated, 22-10 AWG, rings fit 1/4-in bus-bar studs", 8.00, "Amazon", "Amazon TICONN 120-pc kit", 9.99, ""),
 ("S2", "IEC C14 inlet w/ switch + fuse", 1, "2-pk w/ pre-crimped leads, spare incl", 4.00, "Amazon", "Amazon Antrader 2-pk", 9.99, ""),
 ("S2", "IEC cord #2 (plugs into wall inlet)", 1, "second cord of the Amazon Basics 2-pk", 0.00, "Amazon", "Amazon Basics C13 6 ft (2 of 2)", 5.59, ""),
 ("S2", "HUB75 ribbon cables", 0, "CUT from Amazon — each Waveshare panel ships with its own 30 cm 16P cable + power adapter (10 cables > 9 hops needed)", 0.00, "included w/ panels", "", None, ""),
 ("S2", "M3 hardware", 1, "Amazon 750-pc assortment (M3x6-30 + nuts + washers) replaces MC 50-pk $7.99 — frame bolts need nuts, MC pack has none. Adafruit #4207 SMT solder nuts stay out", 6.00, "Amazon", "Amazon Fgruh 750-pc kit", 9.99, ""),
 ("SEC", "TOOLS — buy ONLY what you don't own", None),
 ("TL", "Soldering iron + solder", 1, "Amazon $8.99 kit incl 5 tips + a small solder tube — covers this build AND closes the missing-solder gap; cut MC FNIRSI $39.99 (nicer iron, only worth it if you solder beyond this project). Fabbing the backplane later = buy a real solder spool then", 30.00, "Amazon", "Amazon 60 W adj-temp kit (solder incl)", 8.99, ""),
 ("TL", "Multimeter", 1, "keep ONE of THREE: Amazon $9.98 does 5 V-trim + continuity fine — cut MC MT-1210 $20.49 and Adafruit #850 $24.95", 25.00, "Amazon", "Amazon LJPXHHU DMM", 9.98, ""),
 ("TL", "Wire strippers", 1, "keep ONE of THREE: MC $7.99 covers 10-24 AWG — cut Adafruit #4747 $11.95 and Amazon IRWIN $13.99", 10.00, "Micro Center", "MC SKU 213314", 7.99, "Aug 21"),
 ("TL", "Ratcheting crimper", 1, "BUY THE ~$16 KIND: ratcheting + dies marked INSULATED/HEAT-SHRINK terminals 22-10 AWG (oval red/blue/yellow jaw). Avoid open-barrel or non-insulated dies — they pierce the shrink. Plustool $15.99 or TICONN ~$17 pair with the terminal kit; iCrimp $45 set only adds a ferrule jaw you can skip", 18.00, "Amazon", "", None, ""),
 ("TL", "Precision screwdriver set", 1, "OWNED — skip", 0.00, "owned", "owned", None, ""),
 ("TL", "Heat-shrink kit", 1, "keep ONE of THREE: MC $4.99 — cut Adafruit #4559 $9.95 and Amazon Rindion $4.99 (tie, but MC is same-day). NO heat gun — iron barrel shrinks tubing; cut BOTH heat guns (MC $29.99, Amazon $14.99)", 6.00, "Micro Center", "MC SKU 797464", 4.99, "Aug 21"),
 ("TL", "Jumper wires (M-M / M-F / F-F)", 1, "Amazon 120-pc has all 3 genders (Pico breadboarding) — replaces Adafruit #3141 F-F-only $2.95; MC 40-pin + KS0334 stay cut; elechawk JST kit $19.99 unnecessary (the $0.95 QT cable covers the one Stemma hop)", 5.00, "Amazon", "Amazon 120-pc 20 cm kit", 5.97, ""),
 ("TL", "Micro-USB DATA cable (Pico flashing)", 1, "OWNED — remove Adafruit #2185 $4.95 from that cart", 0.00, "owned", "owned", None, ""),
 ("SEC", "DEFERRED — do NOT buy yet (S3/S6)", None),
 ("DF", "OPTIONAL: 1 domestic panel (Adafruit #3649)", 1, "only if you won't wait 2-4 wk for freight", 55.00, "Adafruit", "", None, ""),
 ("DF", "USB audio line-in (S3 ears)", 1, "Adafruit #1475 $4.95 starts S3; UCA202 only if S4 wants clean stereo — the Amazon $59.90 UCA202+headphones bundle and ATR2XUSB $35.99 are both cut", 30.00, "Adafruit", "Adafruit #1475", 4.95, ""),
 ("DF", "Mini USB mic (S3 room-mic fallback)", 1, "carted THREE times — keep Adafruit $5.95; cut MC $6.99 and Amazon Estiq $5.99", 6.00, "Adafruit", "Adafruit #3367", 5.95, ""),
 ("DF", "VEML7700 + StemmaQT cable (S6)", 1, "auto-brightness; cut Amazon GODIY 2-pk $9.58 dupe", 8.00, "Adafruit", "Adafruit #4162 + #4399", 5.90, ""),
 ("DF", "Diffuser acrylic (S6)", 1, "Amazon 24x24-in 1/4-in translucent white: near-zero trim on a P3/576 mm wall, cut to ~500 mm for P2.5. Thicker than the planned 3 mm opal — stiffer, still diffuses; hold off buying until the pitch decision", 30.00, "Amazon", "Amazon AZM 24x24 sheet", 34.99, ""),
 ("DF", "LED-grey smoked acrylic 3 mm (S6)", 1, "the ND contrast layer — still local plastics shop", 35.00, "local", "", None, ""),
 ("DF", "French cleat + hardwood (S6)", 1, "OOK 533208 12-in/100-lb w/ screws + level; wall is ~15 lb, huge margin. Hardwood frame stock still local", 40.00, "Amazon + local", "Amazon OOK 533208", 15.97, ""),
 ("DF", "Backplane PCB fab run (optional)", 1, "replaces bonnet/bus bars/fuse holders at 9-panel stage; gerbers ready", 130.00, "JLCPCB", "", None, ""),
]

wb = Workbook()
ws = wb.active
ws.title = "Parts Tracker"
ws["A1"] = "ALBUM-ART MATRIX — PARTS TRACKER (minimal build)"
ws["A1"].font = F(bold=True, size=13)
ws["A2"] = ("Yellow = fill as you shop. Found-at/Actual pre-filled from your Adafruit + Micro Center + Amazon carts (2026-08). "
            "Adafruit checkout: code CPDAY = 15% off. Panels + Mean Well trio still open — see PITCH DECISION row and Mouser notes.")
ws["A2"].font = F(italic=True, size=9)

HEADERS = ["Section", "Item", "Qty", "Must-match spec / notes", "Est. $",
           "Suggested source", "Where I found it", "Actual $", "Bought (date)"]
HR = 4
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(HR, c, h)
    cell.font = F(bold=True, color="FFFFFF")
    cell.fill = HEADER
    cell.border = BORDER

TAGLBL = {"S1": "S1", "S2": "S2", "TL": "tool", "DF": "later", "EX": ""}
sec_rows = {"S1": [], "S2": [], "TL": [], "DF": []}
r = HR
for row in ROWS:
    tag = row[0]
    r += 1
    if tag == "SEC":
        ws.cell(r, 1, row[1]).font = F(bold=True)
        for c in range(1, 10):
            ws.cell(r, c).fill = SECTION
            ws.cell(r, c).border = BORDER
        continue
    _, item, qty, spec, est, src, found, actual, bought = row
    vals = [TAGLBL[tag], item, qty, spec, est, src, found, actual, bought]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.border = BORDER
        cell.font = EXAMPLE if tag == "EX" else F()
        cell.alignment = Alignment(wrap_text=(c in (2, 4, 7)), vertical="top")
        if c in (5, 8):
            cell.number_format = MONEY
        if c in (7, 8, 9) and tag != "EX":
            cell.fill = YELLOW
    if tag in sec_rows:
        sec_rows[tag].append(r)

def sf(col, rws):
    return "=SUM(" + ",".join(f"{col}{i}" for i in rws) + ")"

r += 2
for label, keys, bold in (("COMMITTED BUILD (S1 + S2)", ["S1", "S2"], True),
                          ("Tools (if starting from zero)", ["TL"], False),
                          ("Deferred (later stages, incl. optional PCB)", ["DF"], False)):
    cells = [i for k in keys for i in sec_rows[k]]
    ws.cell(r, 2, label).font = F(bold=bold)
    ws.cell(r, 5, sf("E", cells)).number_format = MONEY
    ws.cell(r, 8, sf("H", cells)).number_format = MONEY
    if bold:
        ws.cell(r, 5).font = F(bold=True); ws.cell(r, 8).font = F(bold=True)
    r += 1

for i, w in enumerate([7, 36, 5, 46, 10, 20, 26, 10, 13], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"
wb.save("PARTS-TRACKER.xlsx")
print("regenerated")
